import csv
import re
from pathlib import Path
from openpyxl import Workbook


DEFAULT_PHONE_COLUMNS = [
    "Celular 1",
    "Celular 2",
    "Celular 3",
    "Celular 4",
    "celular 5",
    "Telefone 1",
    "telefone 2",
    "Telefone 3",
    "Telefone 4",
    "telefone 5",
]

# Status em que devemos manter apenas os números associados a eles
STATUS_KEEP_ONLY = {
    "completado - sem interesse",
    "completado - envio de proposta",
    "completado - parente localizado",
    "completado - localizado",
}

# Status em que devemos remover apenas o número associado
STATUS_REMOVE = {
    "completado - ñ conhece",
    "telefone não existe",
    "completado - engano",
}

# Status que devem ir para a aba única "vendidos_sem_contato"
STATUS_VENDIDOS = {
    "completado - vendido",
    "completado - não deseja contato",
}


def normalizar_numero(numero: str) -> str:
    apenas_digitos = re.sub(r"\D", "", numero or "")

    if apenas_digitos.startswith("55") and len(apenas_digitos) > 11:
        apenas_digitos = apenas_digitos[2:]

    if len(apenas_digitos) > 11:
        apenas_digitos = apenas_digitos[-11:]

    return apenas_digitos


def extrair_detalhes(detalhes_brutos: str):
    resultados = []

    for linha in (detalhes_brutos or "").splitlines():
        linha = linha.strip()
        if not linha:
            continue

        padrao = (
            r"^\s*([+\d][\d\s().-]*)\s*-\s*"
            r"\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}\s*-\s*(.+)$"
        )
        match = re.match(padrao, linha)
        if not match:
            continue

        numero, status = match.groups()
        resultados.append((numero.strip(), status.strip()))

    return resultados


def detectar_colunas_telefone(cabecalhos):
    colunas = []
    encontrados = set()

    for nome in cabecalhos or []:
        nome_normalizado = re.sub(r"\s+", "", (nome or "").lower())

        if re.match(r"^(celular|telefone)\d*$", nome_normalizado) and nome_normalizado not in encontrados:
            colunas.append(nome)
            encontrados.add(nome_normalizado)

    return colunas or DEFAULT_PHONE_COLUMNS


def remover_numero_da_linha(linha: dict, numero_alvo: str, colunas_telefone):
    numero_normalizado = normalizar_numero(numero_alvo)

    for coluna in colunas_telefone:
        valor_coluna = linha.get(coluna, "")
        if valor_coluna and normalizar_numero(valor_coluna) == numero_normalizado:
            linha[coluna] = ""


def manter_apenas_numeros(linha: dict, numeros_permitidos: set[str], colunas_telefone):
    numeros_permitidos_norm = {normalizar_numero(n) for n in numeros_permitidos if normalizar_numero(n)}

    for coluna in colunas_telefone:
        valor = linha.get(coluna, "")
        if not valor:
            continue

        valor_norm = normalizar_numero(valor)
        if not valor_norm or valor_norm not in numeros_permitidos_norm:
            linha[coluna] = ""


def classificar_status(status: str) -> str:
    status_normalizado = (status or "").strip().lower()

    if status_normalizado in STATUS_KEEP_ONLY:
        return "manter_apenas_esse"

    if status_normalizado in STATUS_REMOVE:
        return "remover_numero"

    return "manter"


def gerar_caminho_saida_xlsx(caminho_csv: str) -> str:
    caminho = Path(caminho_csv)
    return str(caminho.with_name(f"{caminho.stem}_ajustado.xlsx"))


def processar_csv_para_xlsx(caminho_csv: str, caminho_saida_xlsx: str):
    total_entrada = 0
    total_tratado = 0
    total_por_status = {s: 0 for s in STATUS_KEEP_ONLY}
    total_vendidos_sem_contato = 0

    with open(caminho_csv, "r", encoding="utf-8", newline="") as origem:
        leitor = csv.DictReader(origem, delimiter=";")
        fieldnames = leitor.fieldnames or []
        colunas_telefone = detectar_colunas_telefone(fieldnames)

        wb = Workbook(write_only=True)

        ws_tratado = wb.create_sheet("Tratado")
        ws_tratado.append(fieldnames)

        ws_por_status = {}
        for status in STATUS_KEEP_ONLY:
            ws = wb.create_sheet(status)
            ws.append(fieldnames)
            ws_por_status[status] = ws

        ws_vendidos = wb.create_sheet("vendidos_sem_contato")
        ws_vendidos.append(fieldnames)

        for linha in leitor:
            total_entrada += 1

            status_keep_only_encontrados = set()
            numeros_para_manter_exclusivo = set()
            linha_em_vendidos = False

            for numero, status in extrair_detalhes(linha.get("Detalhes") or ""):
                status_norm = (status or "").strip().lower()
                acao = classificar_status(status)

                if status_norm in STATUS_VENDIDOS:
                    linha_em_vendidos = True

                if acao == "manter_apenas_esse":
                    numeros_para_manter_exclusivo.add(numero)
                    if status_norm in STATUS_KEEP_ONLY:
                        status_keep_only_encontrados.add(status_norm)

                elif acao == "remover_numero":
                    remover_numero_da_linha(linha, numero, colunas_telefone)

            if numeros_para_manter_exclusivo:
                manter_apenas_numeros(linha, numeros_para_manter_exclusivo, colunas_telefone)

            row_out = [linha.get(col, "") for col in fieldnames]

            if linha_em_vendidos:
                ws_vendidos.append(row_out)
                total_vendidos_sem_contato += 1

            elif status_keep_only_encontrados:
                for st in status_keep_only_encontrados:
                    ws_por_status[st].append(row_out)
                    total_por_status[st] += 1

            else:
                ws_tratado.append(row_out)
                total_tratado += 1

        wb.save(caminho_saida_xlsx)

    return total_entrada, total_tratado, total_por_status, total_vendidos_sem_contato


if __name__ == "__main__":
    pasta_execucao = Path.cwd()
    arquivos_csv = sorted(
        arquivo
        for arquivo in pasta_execucao.glob("*.csv")
        if not arquivo.stem.endswith("_ajustado")
    )

    if not arquivos_csv:
        print("Nenhum arquivo .csv encontrado na pasta de execução.")
    else:
        for arquivo in arquivos_csv:
            destino_xlsx = gerar_caminho_saida_xlsx(arquivo)
            entradas, tratadas, por_status, vendidos = processar_csv_para_xlsx(arquivo, destino_xlsx)

            print(f"Arquivo processado: {arquivo.name}")
            print(f"Arquivo gerado:            {Path(destino_xlsx).name}")
            print(f"Linhas originais:         {entradas}")
            print(f"Linhas em 'Tratado':      {tratadas}")
            print(f"Linhas em 'vendidos_sem_contato': {vendidos}")
            print("Linhas copiadas por aba/status (KEEP_ONLY):")
            for st in sorted(por_status.keys()):
                print(f"  - {st}: {por_status[st]}")
            print()
